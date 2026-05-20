"""
Parse Axis Securities PDF/Excel holding statements.
Returns grouped mutual fund data ready for import preview.
"""

import re
from io import BytesIO
from typing import Any


ISIN_PATTERN = re.compile(r"(INF\w{9}|INE\w{9})")
_OCR_JUNK = re.compile(r"[\\|]")  # backslash and pipe are common pdfplumber OCR artifacts


def _clean_stock_name(name: str) -> str:
    name = _OCR_JUNK.sub("", name)
    return re.sub(r"\s+", " ", name).strip()
DATE_PATTERN = re.compile(r"Date\s+(\d{2}-\w{3}-\d{4})")

MONTH_MAP = {
    "Jan": "01", "Feb": "02", "Mar": "03", "Apr": "04",
    "May": "05", "Jun": "06", "Jul": "07", "Aug": "08",
    "Sep": "09", "Oct": "10", "Nov": "11", "Dec": "12",
}


def _parse_report_date(text: str) -> str | None:
    """Extract and convert report date like '08-Mar-2026' → '2026-03-08'."""
    match = DATE_PATTERN.search(text)
    if not match:
        return None
    raw = match.group(1)  # e.g. "08-Mar-2026"
    parts = raw.split("-")
    if len(parts) != 3:
        return None
    day, mon, year = parts
    month_num = MONTH_MAP.get(mon)
    if not month_num:
        return None
    return f"{year}-{month_num}-{day}"


def _parse_holding_lines(full_text: str) -> dict:
    """
    Parse all holding rows from extracted PDF/Excel text.
    Returns grouped dict: { isin: { fund_name, isin, transactions: [...] } }
    """
    funds: dict[str, Any] = {}

    for line in full_text.splitlines():
        line = line.strip()
        if not line:
            continue

        # Try to find ISIN in the line — it anchors our parse
        isin_match = ISIN_PATTERN.search(line)
        if not isin_match:
            continue

        isin = isin_match.group(1)
        isin_start = isin_match.start()
        isin_end = isin_match.end()

        fund_name = line[:isin_start].strip()
        remainder = line[isin_end:].strip()

        # remainder should be: {open_qty} {market_price} {market_value} {investment_amount} {avg_cost} ...
        tokens = remainder.split()
        if len(tokens) < 5:
            continue

        if not isin.startswith("INF"):  # INF = mutual fund ISIN prefix (SEBI standard)
            continue

        numeric_tokens = tokens[:7]
        try:
            open_qty = float(numeric_tokens[0].replace(",", ""))
            market_price = float(numeric_tokens[1].replace(",", ""))
            market_value = float(numeric_tokens[2].replace(",", ""))
            investment_amount = float(numeric_tokens[3].replace(",", ""))
            avg_cost = float(numeric_tokens[4].replace(",", ""))
            # unrealized_pnl = tokens[5] — not used
        except ValueError:
            continue

        if not fund_name:
            continue

        transaction = {
            "units": open_qty,
            "avg_cost": avg_cost,
            "investment_amount": investment_amount,
            "market_price": market_price,
        }

        if isin not in funds:
            funds[isin] = {
                "fund_name": fund_name,
                "isin": isin,
                "transactions": [],
            }
        funds[isin]["transactions"].append(transaction)

    return funds


def _parse_stock_lines(full_text: str) -> dict:
    """
    Parse stock (equity) rows from PDF text.
    Returns grouped dict: { isin: { stock_name, isin, shares, avg_cost, investment_amount, market_price } }
    """
    stocks: dict[str, Any] = {}

    for line in full_text.splitlines():
        line = line.strip()
        if not line:
            continue

        isin_match = ISIN_PATTERN.search(line)
        if not isin_match:
            continue

        isin = isin_match.group(1)
        isin_start = isin_match.start()
        isin_end = isin_match.end()

        raw_name = line[:isin_start].strip()
        remainder = line[isin_end:].strip()

        tokens = remainder.split()
        if len(tokens) < 5:
            continue

        if not isin.startswith("INE"):  # INE = equity ISIN prefix (SEBI standard)
            continue

        numeric_tokens = tokens[:7]
        try:
            open_qty = float(numeric_tokens[0].replace(",", ""))
            market_price = float(numeric_tokens[1].replace(",", ""))
            investment_amount = float(numeric_tokens[3].replace(",", ""))
            avg_cost = float(numeric_tokens[4].replace(",", ""))
        except ValueError:
            continue

        name_warning = bool(_OCR_JUNK.search(raw_name))
        stock_name = _clean_stock_name(raw_name)

        if not stock_name:
            continue

        if isin not in stocks:
            stocks[isin] = {
                "stock_name": stock_name,
                "isin": isin,
                "shares": open_qty,
                "avg_cost": avg_cost,
                "investment_amount": investment_amount,
                "market_price": market_price,
                "name_warning": name_warning,
            }

    return stocks


def parse_stocks_from_pdf(file_bytes: bytes) -> dict:
    """Parse stock holdings from an Axis Securities PDF. Returns { "report_date": ..., "stocks": {...} }"""
    try:
        import pdfplumber
    except ImportError as e:
        raise RuntimeError("pdfplumber is not installed") from e

    full_text_parts = []
    with pdfplumber.open(BytesIO(file_bytes)) as pdf:
        for page in pdf.pages:
            text = page.extract_text()
            if text:
                full_text_parts.append(text)

    full_text = "\n".join(full_text_parts)
    return {"report_date": _parse_report_date(full_text), "stocks": _parse_stock_lines(full_text)}


def parse_stocks_from_excel(file_bytes: bytes) -> dict:
    """Parse stock holdings from an Axis Securities Excel file. Returns { "report_date": ..., "stocks": {...} }"""
    try:
        import openpyxl
    except ImportError as e:
        raise RuntimeError("openpyxl is not installed") from e

    wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))

    header_row_idx = None
    col_map: dict[str, int] = {}
    for i, row in enumerate(rows):
        cells = [str(c).strip() if c is not None else "" for c in row]
        if "ISIN" in cells or "Stock Name" in cells:
            header_row_idx = i
            for j, cell in enumerate(cells):
                col_map[cell] = j
            break

    if header_row_idx is None:
        return {"report_date": None, "stocks": {}}

    header_text = "\n".join(
        " ".join(str(c) if c is not None else "" for c in row)
        for row in rows[:header_row_idx]
    )

    def col(name: str) -> int | None:
        return col_map.get(name)

    name_col = col("Stock Name") or col("Fund Name") or col("Scheme Name")
    isin_col = col("ISIN")
    qty_col = col("Open Qty") or col("Quantity")
    price_col = col("Market Price")
    inv_col = col("Investment Amount")
    avgcost_col = col("Avg Cost") or col("Avg. Cost")
    asset_col = col("Asset Type")

    if isin_col is None:
        return {"report_date": _parse_report_date(header_text), "stocks": {}}

    stocks: dict[str, Any] = {}

    for row in rows[header_row_idx + 1:]:
        if all(c is None for c in row):
            continue

        def get(idx):
            if idx is None or idx >= len(row):
                return None
            return row[idx]

        isin = str(get(isin_col) or "").strip()
        if not ISIN_PATTERN.match(isin):
            continue
        if not isin.startswith("INE"):  # INE = equity ISIN prefix
            continue

        stock_name = str(get(name_col) or "").strip()

        try:
            open_qty = float(str(get(qty_col) or "0").replace(",", ""))
            market_price = float(str(get(price_col) or "0").replace(",", ""))
            investment_amount = float(str(get(inv_col) or "0").replace(",", ""))
            avg_cost = float(str(get(avgcost_col) or "0").replace(",", ""))
        except (ValueError, TypeError):
            continue

        if isin not in stocks:
            stocks[isin] = {
                "stock_name": stock_name,
                "isin": isin,
                "shares": open_qty,
                "avg_cost": avg_cost,
                "investment_amount": investment_amount,
                "market_price": market_price,
                "name_warning": False,  # Excel cells don't have OCR artifacts
            }

    return {"report_date": _parse_report_date(header_text), "stocks": stocks}


def parse_pdf(file_bytes: bytes) -> dict:
    """
    Parse an Axis Securities PDF holding statement.
    Returns: { "report_date": "2026-03-08", "funds": { isin: {...} } }
    """
    try:
        import pdfplumber
    except ImportError as e:
        raise RuntimeError("pdfplumber is not installed") from e

    full_text_parts = []
    with pdfplumber.open(BytesIO(file_bytes)) as pdf:
        for page in pdf.pages:
            text = page.extract_text()
            if text:
                full_text_parts.append(text)

    full_text = "\n".join(full_text_parts)
    report_date = _parse_report_date(full_text)
    funds = _parse_holding_lines(full_text)

    return {"report_date": report_date, "funds": funds}


def parse_excel(file_bytes: bytes) -> dict:
    """
    Parse an Axis Securities Excel holding statement.
    Returns: { "report_date": None, "funds": { isin: {...} } }
    """
    try:
        import openpyxl
    except ImportError as e:
        raise RuntimeError("openpyxl is not installed") from e

    wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))

    # Find header row
    header_row_idx = None
    col_map: dict[str, int] = {}
    for i, row in enumerate(rows):
        cells = [str(c).strip() if c is not None else "" for c in row]
        if "ISIN" in cells or "Stock Name" in cells:
            header_row_idx = i
            for j, cell in enumerate(cells):
                col_map[cell] = j
            break

    if header_row_idx is None:
        return {"report_date": None, "funds": {}}

    # Try to find report date in rows above the header
    report_date = None
    header_text = "\n".join(
        " ".join(str(c) if c is not None else "" for c in row)
        for row in rows[:header_row_idx]
    )
    report_date = _parse_report_date(header_text)

    # Column indices
    def col(name: str) -> int | None:
        return col_map.get(name)

    name_col = col("Stock Name") or col("Fund Name") or col("Scheme Name")
    isin_col = col("ISIN")
    qty_col = col("Open Qty") or col("Quantity")
    price_col = col("Market Price")
    mval_col = col("Market Value")
    inv_col = col("Investment Amount")
    avgcost_col = col("Avg Cost") or col("Avg. Cost")
    asset_col = col("Asset Type")

    if isin_col is None:
        return {"report_date": report_date, "funds": {}}

    funds: dict[str, Any] = {}

    for row in rows[header_row_idx + 1:]:
        if all(c is None for c in row):
            continue

        def get(idx):
            if idx is None or idx >= len(row):
                return None
            return row[idx]

        isin = str(get(isin_col) or "").strip()
        if not ISIN_PATTERN.match(isin):
            continue
        if not isin.startswith("INF"):  # INF = mutual fund ISIN prefix
            continue

        fund_name = str(get(name_col) or "").strip()

        try:
            open_qty = float(str(get(qty_col) or "0").replace(",", ""))
            market_price = float(str(get(price_col) or "0").replace(",", ""))
            investment_amount = float(str(get(inv_col) or "0").replace(",", ""))
            avg_cost = float(str(get(avgcost_col) or "0").replace(",", ""))
        except (ValueError, TypeError):
            continue

        transaction = {
            "units": open_qty,
            "avg_cost": avg_cost,
            "investment_amount": investment_amount,
            "market_price": market_price,
        }

        if isin not in funds:
            funds[isin] = {
                "fund_name": fund_name,
                "isin": isin,
                "transactions": [],
            }
        funds[isin]["transactions"].append(transaction)

    return {"report_date": report_date, "funds": funds}
